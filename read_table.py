from openpyxl import load_workbook
import csv


def save_to_csv(file_name, rows):
    with open(file_name, mode="w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        if rows and not isinstance(rows[0], list):
            writer.writerows([[row] for row in rows])
        else:
            writer.writerows(rows)
    print(f"Збережено у файл: {file_name} | Записів: {len(rows)}")


def extract_structure_list(file_path):
    """
    Зчитує першу колонку.
    Додає до результату рядки де Col1 — непорожній текст (не '№ з/п').
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    result_rows = []

    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        first_value = row[0]
        if isinstance(first_value, str):
            first_value = first_value.strip()
            if first_value and first_value != "№ з/п":
                result_rows.append(first_value)

    return result_rows


def has_value(cell_value) -> bool:
    """Перевіряє чи є в клітинці непуста інформація."""
    return cell_value is not None and str(cell_value).strip() != ""


def classify_row(row) -> str:
    """
    Класифікує рядок за трьома станами:

      'write'  — рахуємо + пишемо в CSV
                 Умова: Col7 (ПІБ) має дані
                 (Col1 при цьому може бути int або пустим)

      'count'  — рахуємо, але НЕ пишемо в CSV
                 Умова: Col7 пусто, але Col1 є int АБО Col2 має дані
                 (посада існує, людини немає — тримаємо нумерацію)

      'skip'   — ігноруємо повністю
                 Умова: Col7 пусто, Col1 не int, Col2 пусто
    """
    col1 = row[0].value
    col2 = row[1].value if len(row) > 1 else None
    col7 = row[6].value if len(row) > 6 else None

    col1_is_int   = isinstance(col1, int)
    col2_has_data = has_value(col2)
    col7_has_data = has_value(col7)

    # Col7 є → людина присутня → записуємо
    if col7_has_data:
        return "write"

    # Col7 пусто, але посада існує (є номер або назва посади) → лічимо
    if col1_is_int or col2_has_data:
        if col1 != "№ з/п":
            return "count"

    return "skip"


def extract_staff_list(file_path):
    """
    Зчитує штатний розпис з Excel (колонки 1–11 + has_background).

    Колонки:
      Col1  — порядковий номер (перезаписується лічильником)
      Col2  — назва посади
      Col3  — Номер СОК
      Col4  — Номер ОЦУ
      Col5  — Статус
      Col6  — ПОЗИЦІЯ
      Col7  — ПІБ
      Col8  — ПРИМІТКА
      Col9  — НОМЕР ПРАЦІВНИКА
      Col10 — Дата народження
      Col11 — Номер телефону
      +has_background — колір фону клітинки Col7

    Нумерація:
      - Повністю перезаписується (ігноруємо числа з файлу).
      - 'write' і 'count' збільшують лічильник.
      - В CSV потрапляють тільки рядки 'write'.
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    result_rows = []
    counter = 0

    for row in ws.iter_rows(max_col=11):
        action = classify_row(row)

        if action == "skip":
            continue

        counter += 1

        if action == "count":
            col2_val = row[1].value if len(row) > 1 else None
            print(f"[{counter}] (посада без працівника): {col2_val}")
            continue

        # action == "write" — визначаємо колір фону Col7
        name_cell = row[6] if len(row) > 6 else None
        has_background = False

        if name_cell is not None:
            fill = name_cell.fill
            if fill and fill.patternType == "solid":
                color_index = fill.fgColor.index
                color_rgb   = fill.fgColor.rgb
                if color_index != 0 or (
                    color_rgb
                    and color_rgb != "00000000"
                    and color_rgb != "FFFFFFFF"
                ):
                    has_background = True

        raw_values = [cell.value for cell in row]  # Col1..Col11
        raw_values[0] = counter                     # виправляємо Col1
        raw_values.append(has_background)           # Col12: колір фону
        result_rows.append(raw_values)

        print(f"[{counter}] {raw_values[6]}")

    return result_rows


if __name__ == "__main__":
    source_excel = "full0506.xlsx"

    try:
        print("Читаємо структуру...")
        structure_data = extract_structure_list(source_excel)
        save_to_csv("str_FULL2.csv", structure_data)

        print("\nЧитаємо персонал (це може зайняти час)...")
        staff_data = extract_staff_list(source_excel)
        save_to_csv("all_FULL2.csv", staff_data)

    except FileNotFoundError:
        print(f"Помилка: Файл {source_excel} не знайдено.")
    except Exception as e:
        print(f"Сталася помилка: {e}")
