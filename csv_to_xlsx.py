import csv
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


def is_section_header(row: list[str]) -> bool:
    """
    Чи є рядок заголовком розділу.
    """
    if not row:
        return False
        
    has_first_value = bool(row[0].strip())
    tail_is_empty_or_false = all(
        cell.strip().lower() in ('', 'false') for cell in row[1:]
    )
    
    return has_first_value and tail_is_empty_or_false


def convert_csv_to_formatted_excel(
    csv_path: Path | str, 
    excel_path: Path | str, 
    encoding: str = 'utf-8'
) -> None:
    """
    Читає CSV файл та конвертує його в XLSX.
    Заголовки розділів об'єднуються, стають Bold, 
    вирівнюються по центру та мають жовте тло.
    """
    csv_file = Path(csv_path)
    excel_file = Path(excel_path)
    
    if not csv_file.exists():
        logger.error(f"Файл {csv_file} не знайдено.")
        raise FileNotFoundError(f"Файл {csv_file} не знайдено.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Облік"

    # Ініціалізація стилів
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    try:
        with csv_file.open(mode='r', encoding=encoding) as f:
            reader = csv.reader(f)
            rows = list(reader)

        if not rows:
            logger.warning("CSV файл порожній. Створено порожній Excel файл.")
            wb.save(excel_file)
            return

        max_cols = max(len(row) for row in rows)

        for row_idx, row in enumerate(rows, start=1):
            if is_section_header(row):
                ws.append([row[0]])
                
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_cols)
                
                # Застосовуємо оновлені стилі до головної клітинки об'єднання
                cell = ws.cell(row=row_idx, column=1)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            else:
                ws.append(row)

        wb.save(excel_file)
        logger.info(f"Успішно збережено відформатований файл: {excel_file}")

    except Exception as e:
        logger.exception(f"Помилка під час конвертації: {e}")
        raise

if __name__ == "__main__":
    convert_csv_to_formatted_excel(Path("FULL0805.csv"), Path("outputtest.xlsx"))
