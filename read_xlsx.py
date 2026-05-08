import csv
import re
from openpyxl import load_workbook


def clean_phone(value):
    if value is None or str(value).strip() == "":
        return False
    cleaned = re.sub(r'[\s\-]', '', str(value))
    return cleaned if cleaned else False

def clean_date(value):
    if value is None or str(value).strip() == "":
        return False
    return str(value).strip()

def has_cell_background(cell):
    """Визначає, чи має клітинка колір фону (крім білого та прозорого)."""
    if not cell or not cell.fill or cell.fill.patternType != "solid":
        return "Ok"
    
    color = cell.fill.fgColor
    if color.index == 0 or color.rgb in ["00000000", "FFFFFFFF", None]:
        return "Ok"
    return "PROBLEM"

def process_excel_to_single_csv():
    source_excel = input("Введіть назву Excel файлу (напр. test_info.xlsx): ").strip()
    if not source_excel:
        print("Помилка: Назва вхідного файлу не може бути порожньою.")
        return

    output_csv = input("Назва результуючого CSV файлу [Enter = test.csv]: ").strip()
    if not output_csv:
        output_csv = "test.csv"

    try:
        wb = load_workbook(source_excel, data_only=True)
        ws = wb.active
        
        combined_data = []
        position_counter = 0

        print(f"Обробка: {source_excel} -> {output_csv}")

        for row in ws.iter_rows(max_col=11):
            col1_val = row[0].value
            
            # Заголовок таблиці
            if col1_val == "№ з/п":
                continue
            
            col2_val = row[1].value if len(row) > 1 else None
            col7_cell = row[6] if len(row) > 6 else None
            col7_val = col7_cell.value if col7_cell else None

            # 1. Визначення структурного підрозділу
            if isinstance(col1_val, str) and col1_val.strip() != "":
                # Додаємо назву та 11 значень False для вирівнювання (разом 12 колонок) 
                structure_row = [col1_val.strip()] + [False] * 11
                combined_data.append(structure_row)
                continue

            # 2. Визначення позицій (якщо є ПІБ або номер/назва)
            is_staff = (col7_val is not None and str(col7_val).strip() != "")
            is_vacant_pos = (isinstance(col1_val, int) or (col2_val is not None and str(col2_val).strip() != ""))

            if is_staff or is_vacant_pos:
                position_counter += 1
                row_values = []
                
                for i in range(11):
                    val = row[i].value if i < len(row) else None
                    
                    if i == 9: # Col10 (Дата народження) 
                        row_values.append(clean_date(val))
                    elif i == 10: # Col11 (Номер телефону) 
                        row_values.append(clean_phone(val))
                    else:
                        row_values.append(val if val is not None and str(val).strip() != "" else False)

                # Оновлюємо номер за порядком
                row_values[0] = position_counter

                # 12-те значення: Колір фону ПІБ 
                row_values.append(has_cell_background(col7_cell))

                combined_data.append(row_values)

        with open(output_csv, mode="w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerows(combined_data)

        print(f"\nУспішно! Створено файл '{output_csv}' з {len(combined_data)} рядками.")

    except FileNotFoundError:
        print(f"Помилка: Файл '{source_excel}' не знайдено.")
    except Exception as e:
        print(f"Виникла помилка: {e}")

if __name__ == "__main__":
    process_excel_to_single_csv()
